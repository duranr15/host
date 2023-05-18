from openpyxl import load_workbook
wb = load_workbook("Host.xlsx")
from json import dumps
import json

def icmp():
    d1={}

    sheet = wb["ICMP"]    
    rows = sheet.max_row    
    columns = sheet.max_column  

    # list to store all the rows of excel file as dictionary
    lst = []
    lst1=[]
    for i in range(1, rows):
        row = {}
        row_data_1 = sheet.cell(row=i+1, column=1).value
        row_data_2 = sheet.cell(row=i+1, column=2).value
        row_data_3 = sheet.cell(row=i+1, column=3).value
        row_data_4 = sheet.cell(row=i+1, column=4).value
        row_data_5 = sheet.cell(row=i+1, column=5).value
        row_data_6 = sheet.cell(row=i+1, column=6).value
        row_data_7 = sheet.cell(row=i+1, column=7).value
        

        d1 = {
                "host":row_data_1,
                "name":row_data_2,
                "templates": [{
                        "name":row_data_3
                    } ],            
                "groups": [{
                        "name":row_data_4
                    } ],
                "interfaces": [{
                        "ip":row_data_5,
                        "interface_ref":row_data_6
                    } ],
                "inventory_mode":row_data_7
                }
        lst1.append(d1)

    data={
        "zabbix_export":{
            "version": "6.0",    
            "hosts":lst1
        }
        
        }
    

    with open('data_icmp.json', 'w') as jf: 
        json.dump(data, jf, indent=4)

def snmp():
    d1={}

    sheet = wb["SNMP"]    
    rows = sheet.max_row    
    columns = sheet.max_column  

    # list to store all the rows of excel file as dictionary
    lst = []
    lst1=[]
    for i in range(1, rows):
        row = {}
        row_data_1 = sheet.cell(row=i+1, column=1).value
        row_data_2 = sheet.cell(row=i+1, column=2).value
        row_data_3 = sheet.cell(row=i+1, column=3).value
        row_data_4 = sheet.cell(row=i+1, column=4).value
        row_data_5 = sheet.cell(row=i+1, column=5).value
        row_data_6 = sheet.cell(row=i+1, column=6).value
        row_data_7 = sheet.cell(row=i+1, column=7).value
        row_data_8 = sheet.cell(row=i+1, column=8).value
        row_data_9 = sheet.cell(row=i+1, column=9).value
        row_data_10 = sheet.cell(row=i+1, column=10).value
        row_data_11 = sheet.cell(row=i+1, column=11).value
        row_data_12 = sheet.cell(row=i+1, column=12).value
        row_data_13 = sheet.cell(row=i+1, column=13).value
        row_data_14 = sheet.cell(row=i+1, column=14).value
        row_data_15= sheet.cell(row=i+1, column=15).value
        row_data_16= sheet.cell(row=i+1, column=16).value
        
        

        d1 = {
                "host":row_data_1,
                "name":row_data_2,
                "templates": [{
                        "name":row_data_3
                    } ],            
                "groups": [{
                        "name":row_data_4
                    } ],
                "interfaces": [{
                        "type":row_data_5,
                        "ip":row_data_6,
                        "port":str(row_data_7),
                        "details":{
                            "version": row_data_8,
                            "securityname": row_data_9,
                            "securitylevel": row_data_10,
                            "authprotocol": row_data_11,
                            "authpassphrase": row_data_12,
                            "privprotocol": row_data_13,
                            "privpassphrase": row_data_14
                        },
                        "interface_ref": row_data_15
                    } ],
                "inventory_mode":row_data_16
                }
        lst1.append(d1)

    data={
        "zabbix_export":{
            "version": "6.0",    
            "hosts":lst1
        }
        
        }
    

    with open('data_snmp.json', 'w') as jf: 
        json.dump(data, jf, indent=4)


if __name__ == '__main__':

    print("Seleccione\n")
    input_option=input('ICMP = 1 , SNMP = 2 \n\n')

    if input_option == "1" :
        icmp()
    if input_option == "2" :
        snmp()

    